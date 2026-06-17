#!/usr/bin/env python3
"""Build data/SOSF-Atlas.xlsx (6 tabs) from data/csv/*.csv.

This is the *bootstrap/refresh* file you import into Google Sheets when you
want the Sheet to match the local CSVs (e.g. after a bulk import). Normal
editing happens IN the Sheet; this just seeds it.

Run:  python3 tools/csv-to-xlsx.py     (needs openpyxl)
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(HERE, "data", "csv")
OUT = os.path.join(HERE, "data", "SOSF-Atlas.xlsx")

TABS = [("TV Movies", "tv-movies.csv"), ("Season 1", "season-1.csv"),
        ("Season 2", "season-2.csv"), ("Season 3", "season-3.csv"),
        ("Season 4", "season-4.csv"), ("Season 5", "season-5.csv")]
WIDE = {"Episode Title": 26, "Description": 34, "Address": 30, "Notes": 40,
        "Air Date": 12, "Timestamp": 11, "Then Image": 18, "Now Image": 18, "Title": 34}

wb = Workbook(); wb.remove(wb.active)
HDR_FILL = PatternFill("solid", start_color="2C1A17")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY = Font(name="Arial", size=11)

for tab, fn in TABS:
    rows = list(csv.reader(open(os.path.join(CSV_DIR, fn), encoding="utf-8")))
    ws = wb.create_sheet(tab)
    header = rows[0]
    for r_i, row in enumerate(rows, 1):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            if r_i == 1:
                cell.font = HDR_FONT; cell.fill = HDR_FILL
                cell.alignment = Alignment(vertical="center")
            else:
                cell.font = BODY
                col = header[c_i - 1] if c_i - 1 < len(header) else ""
                if col in ("Air Date", "Episode", "Timestamp"):
                    cell.number_format = "@"
    for c_i, col in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(c_i)].width = WIDE.get(col, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"
    ws.row_dimensions[1].height = 20

wb.save(OUT)
print("Wrote", OUT)
for tab, _ in TABS:
    ws = wb[tab]
    print(f"  tab '{tab}': {ws.max_row - 1} data rows")
